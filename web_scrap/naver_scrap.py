import requests 
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

import time
import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from dotenv import load_dotenv  # 추가된 부분
import openai
from openai import OpenAI


# .env 파일 로드
load_dotenv()

# 환경 변수에서 API 키 가져오기
# api_key = os.getenv("OPENAI_API_KEY")

# OpenAI 클라이언트 초기화
# client = OpenAI(api_key=api_key)

screenshot_dir = 'screenshots'
# def summarize_text(text):
#     # ChatGPT를 사용하여 텍스트 요약 생성
#     response = client.chat.completions.create(
#         model="gpt-3.5-turbo",
#         messages=[
#             # {"role": "system", "content": "You are a helpful assistant."},
#             {"role": "system", "content": "You are in charge of the marketing team of a credit card company, and market sensing is your job."},            
#             {"role": "user", "content": f"다음 {text}을 접속하여 본문을 찾아내서 본문내용을 200자로 요약해줘:"}
#         ],
#         max_tokens=150
#     )
#     return_str = response.choices[0].message.content
#     print(f'summarize_text = [{return_str}]')
#     return return_str
    

# 1. 네이버에서 검색하고 링크 추출
def search_naver(query):
    url = f"https://search.naver.com/search.naver?query={query}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # 검색 결과에서 링크와 제목 추출
    links = []
    for item in soup.select('.total_tit'):  # 네이버 검색 결과 제목 클래스
        title = item.text.strip()
        link = item.find('a')['href']
        html = get_html_contents(link)
        summary = ""
        # print(f"html = [{html[0:200]}]")
        # summary = ""
        # summary = summarize_text(link)
        # print(summary)
        # print(f"title = [{title}]")        
        # print(f"link = [{link}]")        
        # print(f"summary = [{summary}]")        

        links.append({'title': title, 'link': link, 'html':html, 'summary':summary})
    return links

# 2. 웹페이지 캡쳐 및 본문 내용 가져오기

def capture_and_summarize(link_list):
    
    driver = None  # Ensure driver is always defined
    try:
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--headless")  # UI 없이 실행 (서버 환경에서 필수)
        chrome_options.add_argument("--no-sandbox")  # 권한 문제 방지
        chrome_options.add_argument("--disable-dev-shm-usage")  # 메모리 부족 방지
        chrome_options.add_argument("--disable-gpu")  # GPU 사용 안 함
        chrome_options.add_argument("--remote-debugging-port=9222")  # 디버깅 포트 설정

        print(f'driver = [{driver}]')
        # ChromeDriver 자동 다운로드 및 실행
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        
        print(f'after service = [{service}]')
        print(f'after driver = [{driver}]')
        print(os.system("google-chrome --version"))
        
        # Web scraping logic here...

    except Exception as e:
        print(f"WebDriver initialization failed: {e}")  # Log error
        driver = None  # Explicitly set driver to None in case of failure

    finally:
        if driver:  # Only quit if driver was successfully created
            driver.quit()

    # driver_path = "/mount/src/web_scrap/chromedriver"
    # service = ChromeService(executable_path=driver_path)
    # driver = webdriver.Chrome(service=service, options=chrome_options)
    os.system("google-chrome --version")
    results = []

    os.makedirs(screenshot_dir, exist_ok=True)
    create_excel_from_list(link_list)
    
    for i, item in enumerate(link_list, 1):  # 상위 5개 결과만 처리 (필요에 따라 수정 가능)
        title = item['title']
        url = item['link']
        html = item['html']
        summary = item['summary']
        news_content = fetch_content_by_css(html, '.entry-content')
        print(f"title = [{title}]")
        print(f"url = [{url}]")
        print(f"news_content = [{news_content}]")
        # print(f"html = [{html}]")


        try:
            # 페이지 접속
            driver.get(url)

            # 페이지 크기 계산 및 브라우저 창 크기 조정
            page_width = driver.execute_script('return document.body.scrollWidth')
            page_height = driver.execute_script('return document.body.scrollHeight')
            driver.set_window_size(page_width, page_height)

            # 스크린샷 캡처
            # screenshot = driver.save_screenshot("full_page_screenshot.png")

            time.sleep(2)  # 페이지 로딩 대기
            
            # 스크린샷 저장
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            screenshot_path = f"{screenshot_dir}/{i}.png"
            driver.save_screenshot(screenshot_path)
            
            # 본문 내용 가져오기
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            paragraphs = soup.select('p')  # 본문 내용이 p 태그에 있다고 가정
            content = ' '.join([p.text.strip() for p in paragraphs if p.text.strip()])
            
            # 요약 (간단히 앞 200자 정도로 자름)
            summary = content[:200] + '...' if len(content) > 200 else content
            
            results.append({
                'title': title,
                'link': url,
                'summary': summary,
                'screenshot': screenshot_path,
                'news_content': news_content
            })
        except Exception as e:
            print(f"Error processing {url}: {e}")
            results.append({
                'title': title,
                'link': url,
                'summary': '내용을 가져오지 못함',
                'screenshot': '캡쳐 실패',
                'news_content': '본문가져오기 실패'
            })
    
    driver.quit()
    return results

def fetch_content_by_css(html_contents, css_selector):
    """
    URL과 CSS 선택자를 받아 해당 내용을 추출하는 함수
    :param url: 크롤링할 웹페이지 URL (문자열)
    :param css_selector: 추출하고자 하는 CSS 선택자 (문자열)
    :return: 추출된 내용 리스트
    """
    try:
        # HTTP 요청 보내기
        # headers = {
        #     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        # }
        # response = requests.get(url, headers=headers)
        # response.raise_for_status()  # 요청 실패 시 예외 발생
        
        # HTML 파싱
        soup = BeautifulSoup(html_contents, 'html.parser')
        
        # CSS 선택자로 내용 추출
        elements = soup.select(css_selector)
        if not elements:
            print(f"No elements found for selector '{css_selector}'")
            return []
        
        # 추출된 내용 리스트로 반환
        contents = [element.get_text(strip=True) for element in elements]
        return contents
    
    except requests.exceptions.RequestException as e:
        print(f"Error fetching URL: {e}")
        return []
    except Exception as e:
        print(f"Error processing content: {e}")
        return []



# 3. 엑셀로 저장
def save_to_excel(data):
    df = pd.DataFrame(data)
    filename = f"naver_search_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    df.to_excel(filename, index=False)
    print(f"Results saved to {filename}")

def create_excel_from_list(link_list):
# 타이틀 중 가장 긴 텍스트 길이 계산
    max_title_length = max(len(item['title']) for item in link_list) if link_list else 10
    
    # 기본 "요약" 탭 데이터 준비
    summary_data = []
    for i, item in enumerate(link_list, 1):  # 순번은 1부터 시작
        summary_data.append({
            '순번': i,
            '타이틀': item['title'],
            '요약': item['summary'],
            '링크': item['link']
        })
    global summary_filename
    # Excel 파일 생성
    summary_filename = f"summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(summary_filename, engine='openpyxl') as writer:
        # "요약" 탭 작성
        df_summary = pd.DataFrame(summary_data, columns=['순번', '타이틀', '요약', '링크'])
        df_summary.to_excel(writer, sheet_name='요약', index=False)
        
        # 개별 탭 작성 (순번별로)
        for i, item in enumerate(link_list, 1):
            tab_data = [{
                '순번': i,
                '타이틀': item['title'],
                '요약': '',  # 필요 시 요약 추가 가능
                '링크': item['link']
            }]
            df_tab = pd.DataFrame(tab_data, columns=['순번', '타이틀', '요약', '링크'])
            df_tab.to_excel(writer, sheet_name=str(i), index=False)
        
        # openpyxl 워크북 및 워크시트 가져오기
        workbook = writer.book
        
        # 모든 탭에 테두리 및 열 너비 조정 적용
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 테두리 스타일 정의
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 모든 셀에 테두리 적용
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
            
            # 열 너비 조정
            for col in worksheet.columns:
                column_letter = get_column_letter(col[0].column)
                if column_letter == 'B':  # '타이틀' 열 (두 번째 열)
                    worksheet.column_dimensions[column_letter].width = max_title_length + 2
                else:
                    # 다른 열은 내용에 따라 자동 조정
                    max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                    worksheet.column_dimensions[column_letter].width = max(max_length, 10)
    
            print(f'sheet_name = [{sheet_name}]')
                    
            if sheet_name != '요약':
                screenshot_path = f"{screenshot_dir}/{sheet_name}.png"
                if os.path.exists(screenshot_path):
                    img = Image(screenshot_path)
                    # print(f'img = [{img}]')
                    # original_width = img.width
                    # original_height = img.height
                    # print(f"Original size: width={original_width}, height={original_height}")

                    # # 배율 설정 (예: 1.5배)
                    # scale = 1.5
                    # img.width = int(original_width * scale)
                    # img.height = int(original_height * scale)
                    img.width = 300  # 이미지 크기 조정 (필요 시 수정)
                    img.height = 1800
                    worksheet.add_image(img, 'A3')  # E2 셀에 이미지 삽입

    print(f"Excel file saved as {summary_filename}")


def get_html_contents(url):
    # url = 'https://www.example.com'
    print(f"url = [{url}]")
    if not url :
        raise Exception("url is null")
    
    response = requests.get(url)

    if response.status_code == 200:
        html_content = response.text
        # print(html_content)
        return html_content
    else:
        print(f'Failed to retrieve content: {response.status_code}')
        return ""



def download_excel_file(url, filename):
    """
    웹에서 엑셀 파일을 다운로드합니다.

    Parameters:
    - url (str): 엑셀 파일의 URL
    - filename (str): 저장할 파일 이름
    """
    try:
        # URL에서 파일을 다운로드
        response = requests.get(url)
        
        # 성공적으로 다운로드되었는지 확인
        if response.status_code == 200:
            # 파일을 바이너리 모드로 저장
            with open(filename, 'wb') as file:
                file.write(response.content)
            print(f"{filename} 파일이 성공적으로 다운로드되었습니다.")
        else:
            print(f"다운로드 실패: {response.status_code}")
    except Exception as e:
        print(f"오류 발생: {e}")


# 메인 실행
def web_search(query):
    # query = "카드회사마케팅"
    if query == "" :
        query = "카드회사마케팅" 
        
    print(f"Searching for '{query}' on Naver...")
    
    # 네이버 검색
    link_list = search_naver(query)
    print(f"Found {len(link_list)} links.")
    print(f"link_list = [{link_list}]")
    
    # # 각 링크 처리
    results = capture_and_summarize(link_list)
    
    # # 엑셀로 저장
    save_to_excel(results)
    global summary_filename
    print(f'summary_filename = [{summary_filename}]')
    download_excel_file('https://webscrap-egizqrren8hgkdxx5rqwur.streamlit.app/', summary_filename)

if __name__ == "__main__":
    # 필요한 라이브러리 설치 확인
    try:
        web_search("")
    except ImportError as e:
        print(f"Required library not found: {e}")
        print("Please install required libraries using:")
        print("pip install requests beautifulsoup4 selenium pandas openpyxl")